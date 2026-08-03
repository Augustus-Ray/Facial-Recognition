from __future__ import annotations

import tempfile
import unittest
from datetime import datetime, timezone
from pathlib import Path
from unittest.mock import patch

from openpyxl import load_workbook

from attendance_app.database import (
    EncodingRecord,
    database_session,
    initialize_database,
    record_attendance,
    replace_face_encodings,
)
from attendance_app.excel_exporter import export_daily_attendance_files, export_workbook


class ExcelExporterTests(unittest.TestCase):
    def test_workbook_has_daily_attendance_sheet_with_entry_and_exit_columns(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            db_path = Path(temp_dir) / "attendance.db"
            workbook_path = Path(temp_dir) / "attendance.xlsx"
            attendance_dir = Path(temp_dir) / "attendance"
            event_time = datetime(2026, 8, 1, 8, 30, tzinfo=timezone.utc)
            exit_time = datetime(2026, 8, 1, 16, 5, tzinfo=timezone.utc)

            with database_session(db_path) as conn:
                initialize_database(conn)
                replace_face_encodings(
                    conn,
                    [
                        EncodingRecord("EMP001", "Jane Doe", "jane.jpg", [0.1, 0.2]),
                        EncodingRecord("EMP002", "John Smith", "john.jpg", [0.3, 0.4]),
                    ],
                )
                record_attendance(conn, "EMP001", "Jane Doe", "entrance", event_time=event_time)
                record_attendance(conn, "EMP001", "Jane Doe", "exit", event_time=exit_time)

            with patch("attendance_app.excel_exporter.now_local", return_value=event_time):
                export_workbook(db_path, workbook_path, attendance_dir)

            workbook = load_workbook(workbook_path)
            self.assertIn("Employees", workbook.sheetnames)
            self.assertIn("Attendance of 1 August 2026", workbook.sheetnames)
            self.assertNotIn("Daily Attendance", workbook.sheetnames)

            employees_sheet = workbook["Employees"]
            self.assertEqual(
                [cell.value for cell in employees_sheet[1]],
                ["Employee ID", "Name", "Active"],
            )

            attendance_sheet = workbook["Attendance of 1 August 2026"]
            self.assertEqual(
                [cell.value for cell in attendance_sheet[1]],
                ["Employee ID", "Name", "Entry", "Exit", "Status", "Updated At"],
            )
            self.assertEqual(attendance_sheet["A2"].value, "EMP001")
            self.assertEqual(attendance_sheet["C2"].value, "08:30:00")
            self.assertEqual(attendance_sheet["D2"].value, "16:05:00")
            self.assertEqual(attendance_sheet["E2"].value, "Complete")
            self.assertEqual(attendance_sheet["A3"].value, "EMP002")
            self.assertIsNone(attendance_sheet["C3"].value)
            self.assertIsNone(attendance_sheet["D3"].value)
            self.assertEqual(attendance_sheet["E3"].value, "Absent")

            daily_file = attendance_dir / "Attendance of 1 August 2026.xlsx"
            self.assertTrue(daily_file.exists())
            daily_workbook = load_workbook(daily_file)
            daily_sheet = daily_workbook["01 August 2026"]
            self.assertEqual(
                [cell.value for cell in daily_sheet[1]],
                ["Name of Employee", "ID of Employee", "Clocked In Time", "Clocked Out Time"],
            )
            self.assertEqual(daily_sheet["A2"].value, "Jane Doe")
            self.assertEqual(daily_sheet["B2"].value, "EMP001")
            self.assertEqual(daily_sheet["C2"].value, "08:30:00")
            self.assertEqual(daily_sheet["D2"].value, "16:05:00")
            self.assertEqual(daily_sheet["A3"].value, "John Smith")
            self.assertEqual(daily_sheet["B3"].value, "EMP002")
            self.assertIsNone(daily_sheet["C3"].value)
            self.assertIsNone(daily_sheet["D3"].value)

    def test_daily_attendance_files_can_be_exported_directly(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            db_path = Path(temp_dir) / "attendance.db"
            attendance_dir = Path(temp_dir) / "attendance"
            event_time = datetime(2026, 8, 2, 9, 0, tzinfo=timezone.utc)

            with database_session(db_path) as conn:
                initialize_database(conn)
                replace_face_encodings(conn, [EncodingRecord("EMP001", "Jane Doe", "jane.jpg", [0.1, 0.2])])
                record_attendance(conn, "EMP001", "Jane Doe", "entrance", event_time=event_time)

            with patch("attendance_app.excel_exporter.now_local", return_value=event_time):
                written_files = export_daily_attendance_files(db_path, attendance_dir)

            self.assertEqual(written_files, [attendance_dir / "Attendance of 2 August 2026.xlsx"])
            self.assertTrue(written_files[0].exists())


if __name__ == "__main__":
    unittest.main()
