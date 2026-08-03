from __future__ import annotations

from datetime import date, datetime
from pathlib import Path
from typing import Iterable

from attendance_app.database import access_event_rows, attendance_rows, database_session, initialize_database, list_employees
from attendance_app.config import DEFAULT_PATHS, now_local


def export_workbook(
    db_path: str | Path,
    workbook_path: str | Path,
    attendance_dir: str | Path = DEFAULT_PATHS.attendance_dir,
) -> Path:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise RuntimeError("openpyxl is required to write attendance.xlsx. Run: pip install -r requirements.txt") from exc

    workbook_path = Path(workbook_path)
    workbook_path.parent.mkdir(parents=True, exist_ok=True)

    with database_session(db_path) as conn:
        initialize_database(conn)
        employees = list_employees(conn)
        attendance = attendance_rows(conn)
        events = access_event_rows(conn)

    workbook = Workbook()
    employee_sheet = workbook.active
    employee_sheet.title = "Employees"
    _write_sheet(
        employee_sheet,
        ["Employee ID", "Name", "Active"],
        ([row["employee_id"], row["name"], "Yes" if row["active"] else "No"] for row in employees),
    )

    for attendance_date in _attendance_dates(attendance):
        attendance_sheet = workbook.create_sheet(_attendance_sheet_name(attendance_date))
        attendance_by_employee = {
            row["employee_id"]: row
            for row in attendance
            if row["attendance_date"] == attendance_date.isoformat()
        }
        _write_sheet(
            attendance_sheet,
            ["Employee ID", "Name", "Entry", "Exit", "Status", "Updated At"],
            (
                _attendance_employee_row(employee, attendance_by_employee.get(employee["employee_id"]))
                for employee in employees
                if employee["active"]
            ),
        )

    event_sheet = workbook.create_sheet("Access Events")
    _write_sheet(
        event_sheet,
        ["Event Time", "Employee ID", "Name", "Mode", "Status", "Message", "Distance", "Camera", "Source"],
        (
            [
                row["event_time"],
                row["employee_id"],
                row["employee_name"],
                row["mode"],
                row["status"],
                row["message"],
                row["distance"],
                row["camera_index"],
                row["source"],
            ]
            for row in events
        ),
    )

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for sheet in workbook.worksheets:
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
        sheet.freeze_panes = "A2"
        _autosize_columns(sheet, get_column_letter)

    workbook.save(workbook_path)
    export_daily_attendance_files(db_path, attendance_dir)
    return workbook_path


def export_daily_attendance_files(db_path: str | Path, attendance_dir: str | Path) -> list[Path]:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise RuntimeError("openpyxl is required to write attendance files. Run: pip install -r requirements.txt") from exc

    attendance_dir = Path(attendance_dir)
    attendance_dir.mkdir(parents=True, exist_ok=True)

    with database_session(db_path) as conn:
        initialize_database(conn)
        employees = [row for row in list_employees(conn) if row["active"]]
        attendance = attendance_rows(conn)

    written_files: list[Path] = []
    for attendance_date in _attendance_dates(attendance):
        attendance_by_employee = {
            row["employee_id"]: row
            for row in attendance
            if row["attendance_date"] == attendance_date.isoformat()
        }

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = _daily_sheet_title(attendance_date)
        _write_sheet(
            sheet,
            ["Name of Employee", "ID of Employee", "Clocked In Time", "Clocked Out Time"],
            (
                _daily_attendance_employee_row(employee, attendance_by_employee.get(employee["employee_id"]))
                for employee in employees
            ),
        )

        header_fill = PatternFill("solid", fgColor="1F4E78")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
        sheet.freeze_panes = "A2"
        _autosize_columns(sheet, get_column_letter)

        output_path = attendance_dir / f"{_attendance_file_stem(attendance_date)}.xlsx"
        workbook.save(output_path)
        written_files.append(output_path)

    return written_files


def _write_sheet(sheet, headers: list[str], rows: Iterable[list]) -> None:
    sheet.append(headers)
    for row in rows:
        sheet.append(row)


def _time_only(value: str | None) -> str | None:
    if not value:
        return None
    try:
        return datetime.fromisoformat(value).strftime("%H:%M:%S")
    except ValueError:
        return value[-8:]


def _attendance_dates(attendance: list) -> list[date]:
    dates = {now_local().date()}
    for row in attendance:
        dates.add(date.fromisoformat(row["attendance_date"]))
    return sorted(dates, reverse=True)


def _attendance_sheet_name(attendance_date: date) -> str:
    return f"Attendance of {attendance_date.day} {attendance_date.strftime('%B')} {attendance_date.year}"


def _attendance_file_stem(attendance_date: date) -> str:
    return _attendance_sheet_name(attendance_date)


def _daily_sheet_title(attendance_date: date) -> str:
    return attendance_date.strftime("%d %B %Y")


def _attendance_employee_row(employee, attendance_row) -> list:
    if attendance_row is None:
        return [employee["employee_id"], employee["name"], None, None, "Absent", None]

    entry = _time_only(attendance_row["check_in"])
    exit_time = _time_only(attendance_row["check_out"])
    status = _attendance_status(entry, exit_time)
    return [
        employee["employee_id"],
        employee["name"],
        entry,
        exit_time,
        status,
        attendance_row["updated_at"],
    ]


def _attendance_status(entry: str | None, exit_time: str | None) -> str:
    if entry and exit_time:
        return "Complete"
    if entry:
        return "Entry Only"
    if exit_time:
        return "Exit Only"
    return "Absent"


def _daily_attendance_employee_row(employee, attendance_row) -> list:
    if attendance_row is None:
        return [employee["name"], employee["employee_id"], None, None]
    return [
        employee["name"],
        employee["employee_id"],
        _time_only(attendance_row["check_in"]),
        _time_only(attendance_row["check_out"]),
    ]


def _autosize_columns(sheet, get_column_letter) -> None:
    for column in sheet.columns:
        width = 12
        for cell in column:
            if cell.value is not None:
                width = max(width, min(len(str(cell.value)) + 2, 45))
        sheet.column_dimensions[get_column_letter(column[0].column)].width = width
